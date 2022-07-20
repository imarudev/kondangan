import pathlib
import tkinter as tk
from tkinter.constants import END
from tkinter import Menu, messagebox
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font
from threading import Thread
class BeceanApp:
    def window_contact_us(self):
        ContactUs().run()
    def window_cara_penggunaan(self):
        CaraPenggunaan().run()
    def hubungikami(self):
        Thread(target=self.window_contact_us,daemon=True).start()
    def cara_penggunaan(self):
        Thread(target=self.window_cara_penggunaan,daemon=True).start()
    def validate(self, action, index, value_if_allowed,
                       prior_value, text, validation_type, trigger_type, widget_name):
        if value_if_allowed:
            try:
                float(value_if_allowed)
                return True
            except ValueError:
                return False
        else:
            return False
    def submit(self):
        if self.issubmite == False:
            thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
            wb = Workbook()
            ws = wb.active
            ws.title = 'Halaman 1'
            ws.append(['NAMA','ALAMAT','AMPLOP','BAWAAN'])
            ws.column_dimensions['A'].width=30
            ws.column_dimensions['B'].width=55
            ws.column_dimensions['C'].width=13
            ws.column_dimensions['D'].width=35
            cel_A = ws.cell(row=1,column=1)
            cel_B = ws.cell(row=1,column=2)
            cel_C = ws.cell(row=1,column=3)
            cel_D = ws.cell(row=1,column=4)
            centerin = Alignment(horizontal='center',vertical='center')
            tebalin = Font(bold=True)
            cel_A.border=thin_border
            cel_A.alignment=centerin
            cel_A.font=tebalin
            cel_B.border=thin_border
            cel_B.alignment=centerin
            cel_B.font=tebalin
            cel_C.border=thin_border
            cel_C.alignment=centerin
            cel_C.font=tebalin
            cel_D.border=thin_border
            cel_D.alignment=centerin
            cel_D.font=tebalin
            wb.save(self.spotexcelfile)
            wb.close()
            self.issubmite = True
        nama = self.entry1.get()
        alamat = self.entry3.get()
        amplop = f'{self.entry5.get()}.000'
        bawaan = self.entry6.get()
        work = load_workbook(self.spotexcelfile)
        sheet = work['Halaman 1']
        sheet.append([nama,alamat,amplop,bawaan])
        work.save(self.spotexcelfile)
        work.close()
        self.frame2.destroy()
        self.labelframe6.destroy()
        self.labelframe5.destroy()
        self.labelframe3.destroy()
        self.entry1.delete(0,END)
        self.entry1.focus_set()
        self.label2 = tk.Label(self.frame1)
        self.label2.configure(background='#90be6d', font='system', text='BERHASIL')
        self.label2.grid(column='0', row='5')
        self.count=0
        self.issubmite = True
    def hapus(self):
        self.frame2.destroy()
        self.labelframe6.destroy()
        self.labelframe5.destroy()
        self.labelframe3.destroy()
        self.entry1.delete(0,END)
        self.entry1.focus_set()
        if self.issubmite:
                self.label2.destroy()
        self.count = 0
    def enter(self,event):
        if self.count == 0:
            self.labelframe3 = tk.LabelFrame(self.frame1)
            self.entry3 = tk.Entry(self.labelframe3)
            self.entry3.configure(background='#e9edc9', font='system')
            self.entry3.pack(fill='both', side='top')
            self.labelframe3.configure(background='#90be6d', height='200', text='ALAMAT', width='200')
            self.labelframe3.grid(column='0', ipadx='150', padx='20', pady='5', row='1')
            self.entry3.focus_set()
            if self.issubmite:
                self.label2.destroy()
            self.count+=1
        elif self.count == 1:
            self.labelframe5 = tk.LabelFrame(self.frame1)
            iniangka = (self.frame1.register(self.validate),'%d', '%i', '%P', '%s', '%S', '%v', '%V', '%W')
            self.entry5 = tk.Entry(self.labelframe5, validate = 'key', validatecommand = iniangka)
            self.entry5.configure(background='#e9edc9', font='system')
            self.entry5.pack(fill='both', side='top')
            self.labelframe5.configure(background='#90be6d', height='200', text='AMPLOP', width='200')
            self.labelframe5.grid(column='0', ipadx='150', padx='20', pady='5', row='2')
            self.entry5.focus_set()
            self.count+=1
        elif self.count == 2:
            self.labelframe6 = tk.LabelFrame(self.frame1)
            self.entry6 = tk.Entry(self.labelframe6)
            self.entry6.configure(background='#e9edc9', font='system')
            self.entry6.pack(fill='both', side='top')
            self.labelframe6.configure(background='#90be6d', height='200', text='BAWAAN', width='200')
            self.labelframe6.grid(column='0', ipadx='150', padx='20', pady='5', row='3')
            self.entry6.focus_set()
            self.frame2 = tk.Frame(self.frame1)
            self.button2 = tk.Button(self.frame2)
            self.button2.configure(background='#f72585', text='HAPUS', command=self.hapus)
            self.button2.grid(column='1', ipadx='5', ipady='0', padx='100', pady='5', row='0', sticky='se')
            self.button3 = tk.Button(self.frame2)
            self.button3.configure(background='#4895ef', font='system', text='SIMPAN', command=self.submit)
            self.button3.grid(column='0', ipadx='50', ipady='5', padx='20', pady='5', row='0')
            self.frame2.configure(background='#90be6d', height='200', width='200')
            self.frame2.grid(column='0', row='4', sticky='w')
            self.count+=1
        elif self.count == 3:
            self.submit()
    def __init__(self, master=None):
        # build ui
        self.toplevel1 = tk.Tk() if master is None else tk.Toplevel(master)
        self.frame1 = tk.Frame(self.toplevel1)
        self.labelframe1 = tk.LabelFrame(self.frame1)
        self.entry1 = tk.Entry(self.labelframe1)
        self.entry1.configure(background='#e9edc9', font='system')
        self.entry1.pack(fill='both', side='top')
        self.entry1.focus_set()
        self.labelframe1.configure(background='#90be6d', height='200', text='NAMA', width='200')
        self.labelframe1.grid(column='0', ipadx='150', padx='20', pady='5', row='0')
        self.frame1.configure(background='#90be6d', height='200', width='200')
        self.frame1.pack(side='top')
        self.toplevel1.configure(background='#90be6d', height='200', width='200')
        self.toplevel1.title('kondangan')

        # Menu
        self.menubar = Menu(self.toplevel1)
        self.bantuan = Menu(self.menubar, tearoff=0)
        self.bantuan.add_command(label='Cara Penggunaan', command=self.cara_penggunaan)
        self.bantuan.add_separator()
        self.bantuan.add_command(label='Hubungi Kami', command=self.hubungikami)
        self.jendela = Menu(self.menubar, tearoff=0)
        self.jendela.add_command(label='Keluar', command=self.close)
        self.menubar.add_cascade(label='Jendela',menu=self.jendela)
        self.menubar.add_cascade(label='Bantuan',menu=self.bantuan)
        self.toplevel1.config(menu=self.menubar)
        # Main widget
        self.mainwindow = self.toplevel1
        self.mainwindow.bind('<Return>',self.enter)
        self.count = 0
        pathlib.Path('hasil excel').mkdir(parents=True, exist_ok=True)
        now = datetime.now()
        self.spotexcelfile = f"hasil excel/tanggal {now.strftime('%d %B %Y')}, pukul {now.strftime('%H.%M.%S')}.xlsx"
        self.issubmite = False

    def close(self):
        if messagebox.askokcancel("Keluar", "Apakah anda yakin mau keluar aplikasi?"):
            self.mainwindow.destroy()

    def run(self):
        self.mainwindow.protocol("WM_DELETE_WINDOW", self.close)
        self.mainwindow.mainloop()


class ContactUs:
    def __init__(self, master=None):
        # build ui
        self.toplevel2 = tk.Tk() if master is None else tk.Toplevel(master)
        self.frame3 = tk.Frame(self.toplevel2)
        self.label4 = tk.Label(self.frame3)
        self.label4.configure(font='{Arial} 11 {}', justify='left', text='Silahkan hubungi kontak dibawah ini untuk informasi dan pemesanan software\natau aplikasi yang anda butuhkan')
        self.label4.grid(column='0', row='0', sticky='nw')
        self.label5 = tk.Label(self.frame3)
        self.label5.configure(font='{Arial Baltic} 12 {bold}', justify='left', text='HP/WA:\n082334277668')
        self.label5.grid(column='0', row='1', sticky='w')
        self.label6 = tk.Label(self.frame3)
        self.label6.configure(font='{Arial} 10 {}', justify='left', text='Alamat:\nImam Maruf\nRt27/Rw13, Kel.Sukorejo, Kec.Gandusari, Kab.Trenggalek')
        self.label6.grid(column='0', row='2', sticky='w')
        self.frame3.configure(height='200', width='200')
        self.frame3.pack(ipadx='0', ipady='0', side='top')
        self.toplevel2.configure(background='#a2d2ff', height='200', width='200')
        self.toplevel2.title('Hubungi Pengembang')

        # Main widget
        self.mainwindow = self.toplevel2
    
    def run(self):
        self.mainwindow.mainloop()

class CaraPenggunaan:
    def __init__(self, master=None):
        # build ui
        self.toplevel3 = tk.Tk() if master is None else tk.Toplevel(master)
        self.frame4 = tk.Frame(self.toplevel3)
        self.label10 = tk.Label(self.frame4)
        self.label10.configure(justify='left', text='Cara Penggunaan:\n1. Setelah jendela aplikasi terbuka, silahkan masukkan nama pengunjung, lalu tekan ENTER\n2. Akan terbuka kolom baru. Silahkan masukkan alamat pengunjung, lalu tekan ENTER\n3. Setelah terbuka kolom baru, silahkan masukkan nominal amplop, lalu tekan ENTER\n4. Setelah terbuka kolom baru, silahkan masukkan bawaan pengunjung jikalau pengunjung membawa sesuatu.\n5. Tekan ENTER atau klik tombol SIMPAN untuk menyimpan, atau klik tombol HAPUS untuk menghapus\nSelesai\n\nHasil data dalam bentuk Excel\nAnda dapat membukanya dalam folder "hasil excel" yang ada dalam satu direktori aplikasi ini\n\nTerimakasih\nSalam.....')
        self.label10.pack(side='top')
        self.frame4.configure(height='200', width='200')
        self.frame4.pack(side='top')
        self.toplevel3.configure(height='200', width='200')
        self.toplevel3.title('Cara Penggunaan')
        # Main widget
        self.mainwindow = self.toplevel3
    
    def run(self):
        self.mainwindow.mainloop()



if __name__ == '__main__':
    app = BeceanApp()
    app.run()

